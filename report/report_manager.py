"""ReportManager: handling Excel append and PDF generation (best-effort).
Uses openpyxl and reportlab if available.
"""
from __future__ import annotations
import datetime
import os
import logging
from typing import Optional

try:
    import openpyxl
except Exception:
    openpyxl = None

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.pdfgen import canvas
except Exception:
    canvas = None

logger = logging.getLogger(__name__)

EXCEL_FILE = "متابعة_الأساتذة.xlsx"
REPORTS_DIR = "تقارير_الأساتذة"

os.makedirs(REPORTS_DIR, exist_ok=True)


class ReportManager:
    def __init__(self, excel_path: str = EXCEL_FILE):
        self.excel_path = excel_path
        if openpyxl and not os.path.exists(self.excel_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "المتابعة"
            ws.append(["التاريخ", "الأستاذ", "النوع", "المادة", "الساعة", "الملاحظة"])
            wb.save(self.excel_path)

    def append_row_to_excel(self, date_str: str, prof: str, type_str: str, matiere: str, hour_str: str, note: str = "") -> bool:
        if not openpyxl:
            logger.error("openpyxl غير مثبت")
            return False
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            ws.append([date_str, prof, type_str, matiere, hour_str, note])
            wb.save(self.excel_path)
            return True
        except Exception as e:
            logger.exception("خطأ أثناء الكتابة في Excel: %s", e)
            return False

    def generate_pdf_for_prof(self, prof: str, periode: str, matiere: Optional[str] = None, date_filter: Optional[str] = None) -> Optional[str]:
        if canvas is None:
            logger.warning("reportlab غير مثبت؛ لا يمكن توليد PDF")
            return None
        # best-effort: read excel and filter rows
        try:
            import openpyxl as _op
            wb = _op.load_workbook(self.excel_path)
            ws = wb.active
        except Exception:
            logger.exception("خطأ أثناء فتح ملف Excel")
            return None
        today = datetime.date.today()
        filename = os.path.join(REPORTS_DIR, f"{prof}_{periode}.pdf")
        c = canvas.Canvas(filename, pagesize=A4)
        c.setFont("Helvetica-Bold", 14)
        c.drawCentredString(300, 810, f"تقرير {periode} - {prof}")
        c.setFont("Helvetica", 10)
        c.drawString(50, 790, f"المادة: {matiere if matiere else 'جميع المواد'}")
        c.drawString(50, 775, f"تاريخ الطباعة: {today.strftime('%Y-%m-%d')}")
        y = 750
        for row in ws.iter_rows(min_row=2, values_only=True):
            date, row_prof, ttype, row_matiere, hour, note = row
            if row_prof != prof:
                continue
            if date_filter and date != date_filter:
                continue
            text = f"{date} | {ttype} | {row_matiere} | {hour} | {note or ''}"
            c.drawString(50, y, text[:120])
            y -= 12
            if y < 60:
                c.showPage()
                y = 800
        c.save()
        return filename


