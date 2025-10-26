
### FILE: core/data_manager.py
"""DataManager: centralizes timetable data, parsing, and caching.
Keep Arabic UI labels and comments intact.
"""
from __future__ import annotations
import csv
import os
import re
import logging
from typing import Dict, List, Optional, Set, Any

logger = logging.getLogger(__name__)

DEFAULT_COLORS = ["#FFCCCB", "#B2FF66", "#FFD580", "#AED6F1", "#D7BDE2", "#ABEBC6",
                  "#F9E79F", "#F5CBA7", "#A9DFBF", "#F5B7B1", "#85C1E9", "#D6EAF8", "#FADBD8"]


class DataManager:
    """Manage import, in-memory data structures and simple queries.
    This replaces global dictionaries from the original single-file program.
    """

    def __init__(self):
        #
        self.materials_teachers: Dict[str, List[str]] = {}
        self.materials_colors: Dict[str, str] = {}
        self.teachers_subjects: Dict[str, List[str]] = {}
        self.teachers_classes: Dict[str, List[str]] = {}
        self.classes_teachers: Dict[str, List[str]] = {}
        self.classes_timetable: Dict[str, List[Dict[str, Any]]] = {}
        self.timetable_data: Dict[str, List[Dict[str, Any]]] = {}

    # 
    @staticmethod
    def normalize_teacher_name(raw: Optional[str]) -> str:
        if raw is None:
            return ""
        s = str(raw).strip()
        if not s:
            return ""
        s = re.sub(r'\s+', ' ', s)
        parts = re.split(r'\s*(?:,|/|\+|؛|;|\||&| and )\s*', s)
        first = parts[0].strip()
        first = re.sub(r'\s*\d+$', '', first).strip()
        words = first.split()
        return " ".join(words[:2]) if len(words) >= 2 else first

    @staticmethod
    def split_teachers_field(raw: Optional[str]) -> List[str]:
        if raw is None:
            return []
        s = str(raw).strip()
        if not s:
            return []
        parts = re.split(r'\s*(?:,|/|\+|؛|;|\||&| and )\s*', s)
        out: List[str] = []
        seen: Set[str] = set()
        for p in parts:
            n = DataManager.normalize_teacher_name(p)
            if n and n not in seen:
                seen.add(n)
                out.append(n)
        return out

    @staticmethod
    def extract_weekday(day_field: Optional[str]) -> Optional[int]:
        if day_field is None:
            return None
        s = str(day_field).strip().lower()
        if not s:
            return None
        if s.isdigit():
            v = int(s)
            if 1 <= v <= 7:
                return v - 1
            if 0 <= v <= 6:
                return v
        mapping = {
            'الاثنين': 0, 'اثنين': 0, 'الإثنين': 0,
            'الثلاثاء': 1, 'الثلاثاء': 1,
            'الاربعاء': 2, 'الأربعاء': 2,
            'الخميس': 3, 'الجمعة': 4, 'السبت': 5,
            'الاحد': 6,
            'monday': 0, 'tuesday': 1, 'wednesday': 2, 'thursday': 3, 'friday': 4, 'saturday': 5, 'sunday': 6
        }
        if s in mapping:
            return mapping[s]
        for k, v in mapping.items():
            if k in s:
                return v
        return None

    @staticmethod
    def hour_from_field_enhanced(h_field: Optional[str], period_from_day: Optional[str] = None) -> Optional[int]:
        if h_field is None:
            return None
        s = str(h_field).strip()
        if not s:
            return None
        numbers = re.findall(r'\d+', s)
        if not numbers:
            return None
        n = int(numbers[0])
        if period_from_day == "morning":
            if 1 <= n <= 4:
                return 7 + n
            return 8 + (n % 4)
        elif period_from_day == "afternoon":
            if 1 <= n <= 4:
                return 13 + n
            return 14 + (n % 4)
        else:
            if 1 <= n <= 8:
                if n <= 4:
                    return 7 + n
                else:
                    return 9 + n
            if 8 <= n <= 23:
                return n
            return 8 + (n % 8)

    @staticmethod
    def is_real_class(class_name: Optional[str]) -> bool:
        if not class_name:
            return False
        class_str = str(class_name).strip()
        real_classes = {f"{g}M{i}" for g in (1, 2, 3, 4) for i in range(1, 6)}
        # 
        if class_str in real_classes:
            return True
        if class_str.endswith('_G1') or class_str.endswith('_G2'):
            return False
        if '_G1' in class_str or '_G2' in class_str:
            return False
        if 'استدراك' in class_str or '+' in class_str:
            return False
        return True

    @staticmethod
    def extract_main_class(class_name: Optional[str]) -> Optional[str]:
        if not class_name:
            return class_name
        s = str(class_name).strip()
        for suffix in ('_G1', '_G2'):
            if suffix in s:
                return s.split(suffix)[0]
        return s

    def color_for_subject(self, subject: Optional[str]) -> str:
        if not subject:
            return "#FFFFFF"
        if subject not in self.materials_colors:
            self.materials_colors[subject] = DEFAULT_COLORS[len(self.materials_colors) % len(DEFAULT_COLORS)]
        return self.materials_colors[subject]

    # 
    def import_fet_activities_csv_files(self, paths: List[str]) -> bool:
        """Import multiple CSVs and populate all structures.
        Returns True on success; logs issues but keeps best-effort parsing.
        """
        # 
        self.timetable_data.clear()
        self.materials_teachers.clear()
        self.materials_colors.clear()
        self.teachers_subjects.clear()
        self.teachers_classes.clear()
        self.classes_teachers.clear()
        self.classes_timetable.clear()

        total = 0
        problematic_rows = []

        for path in paths:
            if not path or not os.path.exists(path):
                logger.warning("ملف غير موجود: %s", path)
                continue
            with open(path, encoding='utf-8-sig') as f:
                # 
                header = f.readline()
                f.seek(0)
                delimiter = ','
                if '\t' in header:
                    delimiter = '\t'
                elif ';' in header and header.count(';') > header.count(','):
                    delimiter = ';'
                reader = csv.DictReader(f, delimiter=delimiter)

                for row_num, row in enumerate(reader, start=1):
                    act_id = (row.get('Activity Id') or row.get('ActivityId') or row.get('ID') or '')
                    day_raw = row.get('Day') or row.get('day') or row.get('اليوم') or ''
                    hour_raw = row.get('Hour') or row.get('Period') or row.get('الساعة') or ''
                    subject = (row.get('Subject') or row.get('subject') or row.get('المادة') or '').strip()
                    teachers_raw = (row.get('Teachers') or row.get('Teacher') or row.get('الأستاذ') or '').strip()
                    students_set = (row.get('Students Sets') or row.get('Students') or row.get('Classe') or row.get('الصف') or '').strip()
                    room = (row.get('Room') or row.get('Classroom') or row.get('القاعة') or '').strip()

                    if not any([day_raw, hour_raw, teachers_raw]):
                        continue
                    teacher_names = self.split_teachers_field(teachers_raw)
                    if not teacher_names:
                        continue
                    day_str = str(day_raw)
                    period = None
                    if ' ص' in day_str:
                        period = 'morning'
                    elif ' م' in day_str:
                        period = 'afternoon'
                    day_clean = re.sub(r'\s*[مص]\s*$', '', day_str).strip()
                    wd = self.extract_weekday(day_clean)
                    start_hour = self.hour_from_field_enhanced(hour_raw, period)
                    duration = 1
                    if 'Duration' in row and row.get('Duration'):
                        try:
                            duration = int(row.get('Duration'))
                        except Exception:
                            duration = 1
                    if wd is None or start_hour is None:
                        problematic_rows.append((path, row_num, day_raw, hour_raw))

                    for teacher in teacher_names:
                        if subject:
                            self.materials_teachers.setdefault(subject, set()).add(teacher)
                            self.teachers_subjects.setdefault(teacher, set()).add(subject)
                        if students_set:
                            main_class = self.extract_main_class(students_set)
                            if self.is_real_class(main_class):
                                self.teachers_classes.setdefault(teacher, set()).add(main_class)
                                self.classes_teachers.setdefault(main_class, set()).add(teacher)
                        activity = {
                            'weekday': wd,
                            'start_hour': start_hour,
                            'duration': duration,
                            'subject': subject,
                            'room': room,
                            'class': students_set,
                            'activity_id': act_id or None,
                            'source_file': os.path.basename(path),
                            'original_hour_field': hour_raw,
                            'original_day_field': day_raw,
                            'period': period
                        }
                        self.timetable_data.setdefault(teacher, []).append(activity)
                        if students_set:
                            main_class = self.extract_main_class(students_set)
                            if self.is_real_class(main_class):
                                class_activity = activity.copy()
                                class_activity['teacher'] = teacher
                                class_activity['original_class'] = students_set
                                self.classes_timetable.setdefault(main_class, []).append(class_activity)
                        total += 1

        # 
        for subject in list(self.materials_teachers.keys()):
            self.materials_teachers[subject] = sorted(list(self.materials_teachers[subject]))
        for teacher in list(self.teachers_subjects.keys()):
            self.teachers_subjects[teacher] = sorted(list(self.teachers_subjects[teacher]))
        for teacher in list(self.teachers_classes.keys()):
            self.teachers_classes[teacher] = sorted(list(self.teachers_classes[teacher]))
        for class_name in list(self.classes_teachers.keys()):
            self.classes_teachers[class_name] = sorted(list(self.classes_teachers[class_name]))

        # 
        mats = sorted(list(self.materials_teachers.keys()))
        for i, m in enumerate(mats):
            self.materials_colors[m] = DEFAULT_COLORS[i % len(DEFAULT_COLORS)]

        logger.info("Imported %d activities from %d files (%d problematic rows)", total, len(paths), len(problematic_rows))
        return True

    #
    def sessions_for_prof_on_date(self, prof: str, date_obj) -> Optional[List[Dict[str, Any]]]:
        if prof not in self.timetable_data:
            return None
        wd = date_obj.weekday()
        sessions: List[Dict[str, Any]] = []
        for s in self.timetable_data.get(prof, []):
            s_wd = s.get('weekday')
            s_start = s.get('start_hour')
            s_dur = s.get('duration', 1)
            if s_start is None:
                continue
            if s_wd is not None and s_wd != wd:
                continue
            for h in range(s_start, s_start + max(1, s_dur)):
                if 8 <= h <= 20:
                    sessions.append({'start_hour': h, 'subject': s.get('subject', ''), 'room': s.get('room', ''), 'class': s.get('class', '')})
        uniq = {it['start_hour']: it for it in sessions}
        return [uniq[h] for h in sorted(uniq.keys())]


### FILE: utils/helpers.py
"""Small utilities: file helpers, logging setup, and constants."""
from __future__ import annotations
import logging
import os

LOG_FORMAT = '%(asctime)s %(levelname)s %(name)s: %(message)s'


def setup_logging(log_file: Optional[str] = None, level=logging.INFO):
    handlers = [logging.StreamHandler()]
    if log_file:
        os.makedirs(os.path.dirname(log_file), exist_ok=True)
        handlers.append(logging.FileHandler(log_file, encoding='utf-8'))
    logging.basicConfig(level=level, format=LOG_FORMAT, handlers=handlers)


### FILE: report/report_manager.py
"""ReportManager: handling Excel append and PDF generation (best-effort).
Uses openpyxl and reportlab if available.
"""
from __future__ import annotations
import datetime
import os
import logging
from typing import Optional

try:
    import openpyxl
except Exception:
    openpyxl = None

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.pdfgen import canvas
except Exception:
    canvas = None

logger = logging.getLogger(__name__)

EXCEL_FILE = "متابعة_الأساتذة.xlsx"
REPORTS_DIR = "تقارير_الأساتذة"

os.makedirs(REPORTS_DIR, exist_ok=True)


class ReportManager:
    def __init__(self, excel_path: str = EXCEL_FILE):
        self.excel_path = excel_path
        if openpyxl and not os.path.exists(self.excel_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "المتابعة"
            ws.append(["التاريخ", "الأستاذ", "النوع", "المادة", "الساعة", "الملاحظة"])
            wb.save(self.excel_path)

    def append_row_to_excel(self, date_str: str, prof: str, type_str: str, matiere: str, hour_str: str, note: str = "") -> bool:
        if not openpyxl:
            logger.error("openpyxl غير مثبت")
            return False
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            ws.append([date_str, prof, type_str, matiere, hour_str, note])
            wb.save(self.excel_path)
            return True
        except Exception as e:
            logger.exception("خطأ أثناء الكتابة في Excel: %s", e)
            return False

    def generate_pdf_for_prof(self, prof: str, periode: str, matiere: Optional[str] = None, date_filter: Optional[str] = None) -> Optional[str]:
        if canvas is None:
            logger.warning("reportlab غير مثبت؛ لا يمكن توليد PDF")
            return None
        # best-effort: read excel and filter rows
        try:
            import openpyxl as _op
            wb = _op.load_workbook(self.excel_path)
            ws = wb.active
        except Exception:
            logger.exception("خطأ أثناء فتح ملف Excel")
            return None
        today = datetime.date.today()
        filename = os.path.join(REPORTS_DIR, f"{prof}_{periode}.pdf")
        c = canvas.Canvas(filename, pagesize=A4)
        c.setFont("Helvetica-Bold", 14)
        c.drawCentredString(300, 810, f"تقرير {periode} - {prof}")
        c.setFont("Helvetica", 10)
        c.drawString(50, 790, f"المادة: {matiere if matiere else 'جميع المواد'}")
        c.drawString(50, 775, f"تاريخ الطباعة: {today.strftime('%Y-%m-%d')}")
        y = 750
        for row in ws.iter_rows(min_row=2, values_only=True):
            date, row_prof, ttype, row_matiere, hour, note = row
            if row_prof != prof:
                continue
            if date_filter and date != date_filter:
                continue
            text = f"{date} | {ttype} | {row_matiere} | {hour} | {note or ''}"
            c.drawString(50, y, text[:120])
            y -= 12
            if y < 60:
                c.showPage()
                y = 800
        c.save()
        return filename


### FILE: ui/main_ui.py
"""Tkinter main UI entry point. Keeps Arabic labels intact.
This module uses DataManager and ReportManager to provide functionality.
"""
from __future__ import annotations
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import logging
import os
from typing import List
from tkcalendar import DateEntry

from core.data_manager import DataManager
from report.report_manager import ReportManager

logger = logging.getLogger(__name__)

BG = "#f2f7f9"
COLOR_PRESENT = "#A9DFBF"
COLOR_ABSENT = "#F1948A"
COLOR_LATE = "#F5B041"


class UIManager:
    def __init__(self, root: tk.Tk, data_manager: DataManager, report_manager: ReportManager):
        self.root = root
        self.dm = data_manager
        self.rm = report_manager

    def build_main_ui(self):
        for w in self.root.winfo_children():
            w.destroy()
        tk.Label(self.root, text='📚 برنامج منصوري لمتابعة عمل الأساتذة', font=("Arial", 22, "bold"), bg=BG, fg='#004d80').pack(pady=8)
        top_controls = tk.Frame(self.root, bg=BG); top_controls.pack(pady=6)
        ttk.Button(top_controls, text="📅 استيراد جدول CSV", command=self.import_csv_and_refresh).pack(side='left', padx=8)
        ttk.Button(top_controls, text="🔍 التحقق من البيانات", command=self.verify_timetable_match).pack(side='left', padx=8)
        ttk.Button(top_controls, text="🏫 عرض جميع الأقسام", command=self.open_classes_window).pack(side='left', padx=8)
        tk.Label(top_controls, text="(استيراد CSV من FET)", bg=BG).pack(side='left', padx=8)

        stats_frame = tk.Frame(self.root, bg=BG); stats_frame.pack(pady=6)
        tk.Label(stats_frame, text=f"📚 المواد: {len(self.dm.materials_teachers)}", bg=BG, font=("Arial", 11)).pack(side='left', padx=12)
        tk.Label(stats_frame, text=f"👨‍🏫 الأساتذة: {len(self.dm.timetable_data)}", bg=BG, font=("Arial", 11)).pack(side='left', padx=12)
        tk.Label(stats_frame, text=f"🏫 الأقسام: {len(self.dm.classes_timetable)}", bg=BG, font=("Arial", 11)).pack(side='left', padx=12)

        main_frame = tk.Frame(self.root, bg=BG); main_frame.pack(pady=10, fill='both', expand=True)
        left = tk.Frame(main_frame, bg=BG); left.pack(side='left', fill='both', expand=True, padx=8, pady=8)
        if not self.dm.materials_teachers:
            tk.Label(left, text="لم يتم استيراد الجدول بعد. اضغط 'استيراد جدول CSV' أو ضع ملف CSV في المجلد.", bg=BG).pack(pady=12)
        else:
            mats = list(self.dm.materials_teachers.keys())
            cols = 3; r = c = 0
            for i, mat in enumerate(mats):
                color = self.dm.materials_colors.get(mat, '#ddd')
                btn = tk.Button(left, text=mat, bg=color, font=("Arial", 12, "bold"), fg="black", width=26, height=2,
                                relief="raised", bd=2, command=lambda m=mat: self.open_material_window(m))
                btn.grid(row=r, column=c, padx=10, pady=10)
                c += 1
                if c >= cols:
                    c = 0; r += 1
        right = tk.Frame(main_frame, bg=BG, width=360); right.pack(side='right', fill='y', padx=8, pady=8)
        tk.Label(right, text="👨‍🏫 قائمة الأساتذة:", font=("Arial", 12, "bold"), bg=BG).pack(pady=6)
        lb = tk.Listbox(right, width=36, height=20); lb.pack(pady=4)
        for t in sorted(self.dm.timetable_data.keys()): lb.insert('end', t)

        def on_select_teacher(evt=None):
            sel = lb.curselection()
            if not sel: return
            prof = lb.get(sel[0])
            # opens a tracking window (simplified for clarity)
            messagebox.showinfo('معلومة', f'فتح ملف الأستاذ: {prof}')

        lb.bind("<Double-Button-1>", on_select_teacher)
        ttk.Button(right, text="فتح ملف الأستاذ", command=on_select_teacher).pack(pady=6)
        ttk.Button(right, text="استيراد CSV", command=self.import_csv_and_refresh).pack(pady=6)
        ttk.Button(right, text="عرض الأقسام", command=self.open_classes_window).pack(pady=6)

    # ----- simplified windows (you can expand) -----
    def import_csv_and_refresh(self):
        file_paths = filedialog.askopenfilenames(title="استيراد جدول CSV من FET",
                                                 filetypes=[("CSV files", "*.csv"), ("All files", "*.*")])
        if not file_paths:
            return
        ok = self.dm.import_fet_activities_csv_files(list(file_paths))
        if ok:
            self.build_main_ui()

    def verify_timetable_match(self):
        if not self.dm.timetable_data:
            messagebox.showinfo("Info", "لم يتم استيراد أي بيانات بعد")
            return
        summary = f"البيانات المستوردة:\n" \
                  f"عدد الأساتذة: {len(self.dm.timetable_data)}\n" \
                  f"عدد المواد: {len(self.dm.materials_teachers)}\n" \
                  f"عدد الأقسام: {len(self.dm.classes_timetable)}\n"
        messagebox.showinfo("تحقق من البيانات", summary)

    def open_classes_window(self):
        if not self.dm.classes_timetable:
            messagebox.showinfo("معلومة", "لم يتم استيراد أي بيانات عن الأقسام بعد")
            return
        top = tk.Toplevel(self.root); top.title("جميع الأقسام"); top.geometry("500x600"); top.configure(bg=BG)
        tk.Label(top, text="🏫 جميع الأقسام", font=("Arial", 16, "bold"), bg=BG).pack(pady=12)
        lb = tk.Listbox(top, width=60, height=25)
        lb.pack(padx=8, pady=8, fill='both', expand=True)
        for class_name in sorted(self.dm.classes_timetable.keys()):
            teachers_count = len(self.dm.classes_teachers.get(class_name, []))
            activities_count = len(self.dm.classes_timetable.get(class_name, []))
            lb.insert('end', f"{class_name} ({teachers_count} أستاذ - {activities_count} حصة)")

    def open_material_window(self, matiere: str):
        top = tk.Toplevel(self.root); top.title(f"أساتذة {matiere}"); top.geometry("380x480"); top.configure(bg=BG)
        tk.Label(top, text=f"أساتذة {matiere}", font=("Arial", 14, "bold"), bg=BG).pack(pady=10)
        profs = self.dm.materials_teachers.get(matiere, [])
        if not profs:
            tk.Label(top, text="لا يوجد أساتذة مسجلين لهذه المادة", bg=BG).pack(pady=8); return
        for p in profs:
            ttk.Button(top, text=p, width=34, command=lambda pr=p, m=matiere: messagebox.showinfo('معلومة', f'فتح ملف {pr}')).pack(pady=6)


### FILE: run.py
"""Entry point to start the refactored app."""
from __future__ import annotations
import tkinter as tk
from utils.helpers import setup_logging
from core.data_manager import DataManager
from report.report_manager import ReportManager
from ui.main_ui import UIManager


def main():
    setup_logging(None)
    dm = DataManager()
    rm = ReportManager()
    root = tk.Tk()
    root.title('ناظر المدرسة - Suivi des enseignants')
    root.geometry('1120x760')
    root.configure(bg='#f2f7f9')
    ui = UIManager(root, dm, rm)
    ui.build_main_ui()
    root.mainloop()

if __name__ == '__main__':
    main()

### FILE: tests/test_importer.py
"""Simple smoke test for importer. Run with pytest after saving a sample CSV file."""
from core.data_manager import DataManager
import tempfile

SAMPLE_CSV = '''Activity Id,Day,Hour,Subject,Teachers,Room,Students Sets
1,الاثنين,1,Math,Ali Ahmed,101,4M1
2,الثلاثاء,2,Physics,Mohamed Salah,102,4M2
'''


def test_import_sample(tmp_path):
    p = tmp_path / "sample.csv"
    p.write_text(SAMPLE_CSV, encoding='utf-8')
    dm = DataManager()
    ok = dm.import_fet_activities_csv_files([str(p)])
    assert ok
    assert 'Math' in dm.materials_teachers
    assert 'Ali Ahmed' in dm.timetable_data

# End of project content
