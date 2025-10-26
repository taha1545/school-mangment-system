import os
import io
import csv
import re
import glob
import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from tkcalendar import DateEntry

# مكتبات اختيارية للمعاينة
try:
    import fitz
    from PIL import Image, ImageTk

    FITZ_AVAILABLE = True
except Exception:
    FITZ_AVAILABLE = False

# reportlab لتصدير PDF
try:
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors

    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False

# ملفات ومجلدات
CSV_CANDIDATES = ["activities.csv", "fet_timetable_full.csv", "مقترح10_timetable.csv", "timetable.csv"]
EXCEL_FILE = "متابعة_الأساتذة.xlsx"
REPORTS_DIR = "تقارير_الأساتذة"
os.makedirs(REPORTS_DIR, exist_ok=True)

# أنشئ ملف المتابعة إن لم يكن موجودًا
if not os.path.exists(EXCEL_FILE):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "المتابعة"
    ws.append(["التاريخ", "الأستاذ", "النوع", "المادة", "الساعة", "الملاحظة"])
    wb.save(EXCEL_FILE)

# ألوان وواجهة
BG = "#f2f7f9"
COLOR_PRESENT = "#A9DFBF"
COLOR_ABSENT = "#F1948A"
COLOR_LATE = "#F5B041"
DEFAULT_COLORS = ["#FFCCCB", "#B2FF66", "#FFD580", "#AED6F1", "#D7BDE2", "#ABEBC6",
                  "#F9E79F", "#F5CBA7", "#A9DFBF", "#F5B7B1", "#85C1E9", "#D6EAF8", "#FADBD8"]

# هياكل البيانات
materials_teachers = {} 
materials_colors = {}
teachers_subjects = {} 
teachers_classes = {}  
classes_teachers = {}  
classes_timetable = {} 
timetable_data = {}  


# ---------------- دوال مساعدة محسنة ----------------
def normalize_teacher_name(raw):
    """Normalize teacher name: take first part, keep up to two words."""
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s:
        return ""
    s = re.sub(r'\s+', ' ', s)
    # 
    parts = re.split(r'\s*(?:,|/|\+|؛|;|\||&| and )\s*', s)
    first = parts[0].strip()
    first = re.sub(r'\s*\d+$', '', first).strip()
    words = first.split()
    return " ".join(words[:2]) if len(words) >= 2 else first


def split_teachers_field(raw):
    """Split a Teachers field into individual normalized names, deduped."""
    if raw is None:
        return []
    s = str(raw).strip()
    if not s:
        return []
    parts = re.split(r'\s*(?:,|/|\+|؛|;|\||&| and )\s*', s)
    out = []
    seen = set()
    for p in parts:
        n = normalize_teacher_name(p)
        if n and n not in seen:
            seen.add(n)
            out.append(n)
    return out


def extract_weekday(day_field):
    """
    Convert Day field to weekday index 0=Mon .. 6=Sun.
    Enhanced with more comprehensive mapping.
    """
    if day_field is None:
        return None

    s = str(day_field).strip().lower()
    if not s:
        return None

    # Handle numeric values (FET often uses 1-7 for Monday-Sunday)
    if s.isdigit():
        v = int(s)
        if 1 <= v <= 7:
            return v - 1 
        if 0 <= v <= 6:
            return v

    # Comprehensive mapping
    mapping = {
        # Arabic
        'الاثنين': 0, 'اثنين': 0, 'الإثنين': 0, 'أول': 0, '1': 0,
        'الثلاثاء': 1, 'الثلثاء': 1, 'ثلاثاء': 1, 'ثاني': 1, '2': 1,
        'الاربعاء': 2, 'الأربعاء': 2, 'اربعاء': 2, 'أربعاء': 2, 'ثالث': 2, '3': 2,
        'الخميس': 3, 'خميس': 3, 'رابع': 3, '4': 3,
        'الجمعة': 4, 'جمعة': 4, 'خامس': 4, '5': 4,
        'السبت': 5, 'سبت': 5, 'سادس': 5, '6': 5,
        'الاحد': 6, 'الأحد': 6, 'احد': 6, 'أحد': 6, 'سابع': 6, '7': 6,

        # English
        'monday': 0, 'mon': 0, 'mo': 0,
        'tuesday': 1, 'tue': 1, 'tu': 1,
        'wednesday': 2, 'wed': 2, 'we': 2,
        'thursday': 3, 'thu': 3, 'th': 3,
        'friday': 4, 'fri': 4, 'fr': 4,
        'saturday': 5, 'sat': 5, 'sa': 5,
        'sunday': 6, 'sun': 6, 'su': 6,

        # French
        'lundi': 0, 'lu': 0,
        'mardi': 1, 'ma': 1,
        'mercredi': 2, 'me': 2,
        'jeudi': 3, 'je': 3,
        'vendredi': 4, 've': 4,
        'samedi': 5, 'sa': 5,
        'dimanche': 6, 'di': 6
    }

    s_lower = s.lower()

    # Check for exact matches first
    for key, value in mapping.items():
        if key == s_lower:
            return value

    # Check for partial matches
    for key, value in mapping.items():
        if key in s_lower:
            return value

    return None


def hour_from_field_enhanced(h_field, period_from_day=None):
    """
    نسخة محسنة تأخذ بعين الاعتبار الفترة من حقل اليوم
    """
    if h_field is None:
        return None

    s = str(h_field).strip()
    if not s:
        return None

    # استخراج رقم الحصة
    numbers = re.findall(r'\d+', s)
    if not numbers:
        return None

    n = int(numbers[0])

    # استخدام المعلومات من حقل اليوم إذا كانت متوفرة
    if period_from_day == "morning":
        if 1 <= n <= 4:
            return 7 + n 
        else:
            return 8 + (n % 4)

    elif period_from_day == "afternoon":
        if 1 <= n <= 4:
            return 13 + n 
        else:
            return 14 + (n % 4)

    else:
        # النظام الافتراضي إذا لم تكن هناك معلومات عن الفترة
        if 1 <= n <= 8:
            if n <= 4:
                return 7 + n  
            else:
                return 9 + n  
        elif 8 <= n <= 23:
            return n
        else:
            return 8 + (n % 8)


def hour_from_field(h_field):
    """
    Convert Hour/Period field to an actual hour integer.
    COMPLETELY REWRITTEN for proper FET period handling
    """
    return hour_from_field_enhanced(h_field)


def is_real_class(class_name):
    """
    تحديد إذا كان اسم القسم حقيقيًا أو مجرد مجموعة (G1, G2)
    """
    if not class_name:
        return False

    class_str = str(class_name).strip()

    # الأقسام الحقيقية (عدل هذه القائمة حسب أقسامك الفعلية)
    real_classes = [
        '4M1', '4M2', '4M3', '4M4', '4M5',
        '3M1', '3M2', '3M3', '3M4', '3M5',
        '2M1', '2M2', '2M3', '2M4', '2M5',
        '1M1', '1M2', '1M3', '1M4', '1M5'
    ]

    # إذا كان القسم موجودًا في القائمة فهو حقيقي
    if class_str in real_classes:
        return True

    # إذا انتهى بـ _G1 أو _G2 فهو مجموعة وليس قسم حقيقي
    if class_str.endswith('_G1') or class_str.endswith('_G2'):
        return False

    # إذا كان يحتوي على "G1" أو "G2" فهو مجموعة
    if '_G1' in class_str or '_G2' in class_str:
        return False

    # استبعاد الأنشطة الخاصة مثل "استدراك"
    if 'استدراك' in class_str or '+' in class_str:
        return False

    return True


def extract_main_class(class_name):
    """
    استخراج القسم الرئيسي من اسم المجموعة
    مثال: "4M1_G1" -> "4M1"
    """
    if not class_name:
        return class_name

    class_str = str(class_name).strip()

    # إزالة _G1, _G2
    if '_G1' in class_str:
        return class_str.split('_G1')[0]
    elif '_G2' in class_str:
        return class_str.split('_G2')[0]

    return class_str


def color_for_subject(subject):
    """Assign a color for a subject (persistent during runtime)."""
    if not subject:
        return "#FFFFFF"
    if subject not in materials_colors:
        materials_colors[subject] = DEFAULT_COLORS[len(materials_colors) % len(DEFAULT_COLORS)]
    return materials_colors[subject]


# ---------------- دوال استيراد CSV محسنة ----------------
def import_fet_activities_csv_files(paths):
    """
    Import multiple CSV files produced by FET (Activities export).
    - Improved parsing for different FET export formats
    - Better handling of multiple teachers per activity
    - Enhanced time slot detection with proper afternoon periods
    - Added class timetable tracking
    - Filter out groups (G1, G2) and keep only real classes
    """
    global timetable_data, materials_teachers, teachers_subjects, teachers_classes, classes_teachers, classes_timetable, materials_colors

    # Clear all data structures
    timetable_data.clear()
    materials_teachers.clear()
    teachers_subjects.clear()
    teachers_classes.clear()
    classes_teachers.clear()
    classes_timetable.clear()
    materials_colors.clear()

    seen_ids = set()
    seen_hashes = set()
    total = 0
    problematic_rows = []

    try:
        for path in paths:
            if not path or not os.path.exists(path):
                continue

            print(f"📁 جاري معالجة الملف: {path}")

            with open(path, encoding='utf-8-sig') as f:
                # Detect delimiter and encoding
                sample = f.readline()
                print(f"📄 عينة من الملف: {sample}")
                f.seek(0)

                # Try different delimiters
                reader = None
                for delimiter in [',', ';', '\t']:
                    try:
                        f.seek(0)
                        reader = csv.DictReader(f, delimiter=delimiter)
                        # Test reading first row
                        first_row = next(reader)
                        f.seek(0)
                        reader = csv.DictReader(f, delimiter=delimiter)
                        print(f"✅ تم استخدام delimiter: {delimiter}")
                        break
                    except Exception as e:
                        continue

                if reader is None:
                    f.seek(0)
                    reader = csv.DictReader(f)  # fallback to default
                    print("⚠️ استخدام delimiter افتراضي")

                for row_num, row in enumerate(reader):
                    # Enhanced field detection for different FET formats
                    act_id = (row.get('Activity Id') or row.get('ActivityId') or
                              row.get('ID') or row.get('id') or row.get('Numéro') or "").strip()

                    # Try multiple possible column names for day
                    day_raw = None
                    for day_col in ['Day', 'day', 'Jour', 'jour', 'اليوم', 'JOUR', 'يوم']:
                        if day_col in row and row[day_col]:
                            day_raw = row[day_col]
                            break
                    if day_raw is None:
                        day_raw = ""

                    # Try multiple possible column names for hour
                    hour_raw = None
                    for hour_col in ['Hour', 'hour', 'Period', 'Heure', 'Start time', 'Start_time',
                                     'الساعة', 'الحصة', 'HOUR', 'Période']:
                        if hour_col in row and row[hour_col]:
                            hour_raw = row[hour_col]
                            break
                    if hour_raw is None:
                        hour_raw = ""

                    subject = (row.get('Subject') or row.get('subject') or
                               row.get('Matière') or row.get('Course') or
                               row.get('المادة') or row.get('SUBJECT') or "").strip()

                    teachers_raw = (row.get('Teachers') or row.get('Teacher') or
                                    row.get('teacher') or row.get('Enseignant') or
                                    row.get('الأستاذ') or row.get('TEACHERS') or "").strip()

                    room = (row.get('Room') or row.get('room') or row.get('Salle') or
                            row.get('Classroom') or row.get('Local') or row.get('القاعة') or row.get(
                                'ROOM') or "").strip()

                    students_set = (row.get('Students Sets') or row.get('Student Sets') or
                                    row.get('Students') or row.get('Classe') or
                                    row.get('الصف') or row.get('STUDENTS') or row.get('Class') or "").strip()

                    print(f"  📊 الصف {row_num}:")
                    print(f"    📅 اليوم: '{day_raw}'")
                    print(f"    ⏰ الساعة: '{hour_raw}'")
                    print(f"    📚 المادة: '{subject}'")
                    print(f"    👨‍🏫 الأساتذة: '{teachers_raw}'")
                    print(f"    🏫 القسم: '{students_set}'")

                    # Skip empty rows
                    if not any([day_raw, hour_raw, teachers_raw]):
                        print("    ⏭️ تخطي الصف - فارغ")
                        continue

                    teacher_names = split_teachers_field(teachers_raw)
                    if not teacher_names:
                        print("    ⏭️ تخطي - لا يوجد أساتذة")
                        continue

                    # إصلاح مؤقت للتعرف على الفترات من حقل اليوم
                    day_str = str(day_raw)
                    if " ص" in day_str:
                        period = "morning"
                    elif " م" in day_str:
                        period = "afternoon"
                    else:
                        period = None

                    # استخراج اليوم بعد إزالة الفترة
                    day_clean = re.sub(r'\s*[مص]\s*$', '', day_str).strip()
                    day_clean = re.sub(r'^\d+\s*', '', day_clean).strip()

                    wd = extract_weekday(day_clean)
                    start_hour = hour_from_field_enhanced(hour_raw, period)

                    print(f"    🔄 اليوم المحول: {wd} (الفترة: {period})")
                    print(f"    🔄 الساعة المحولة: {start_hour}")

                    # Track problematic rows
                    if wd is None or start_hour is None:
                        problematic_rows.append({
                            'row': row_num,
                            'day_raw': day_raw,
                            'hour_raw': hour_raw,
                            'day_converted': wd,
                            'hour_converted': start_hour
                        })
                        print("    ⚠️ تحذير: يوم أو ساعة غير محدد")

                    # Enhanced duration detection
                    duration = 1
                    if 'Duration' in row and row['Duration']:
                        try:
                            duration = int(row['Duration'])
                        except:
                            duration = 1

                    for teacher in teacher_names:
                        # Add to subject->teachers mapping
                        if subject:
                            materials_teachers.setdefault(subject, set()).add(teacher)
                            teachers_subjects.setdefault(teacher, set()).add(subject)

                        # Add to teacher->classes mapping (فقط الأقسام الحقيقية)
                        if students_set:
                            main_class = extract_main_class(students_set)
                            if is_real_class(main_class):
                                teachers_classes.setdefault(teacher, set()).add(main_class)
                                # Add to class->teachers mapping
                                classes_teachers.setdefault(main_class, set()).add(teacher)
                                print(f"    ✅ إضافة قسم حقيقي: {main_class}")

                        # Create activity entry for teacher
                        activity = {
                            'weekday': wd,
                            'start_hour': start_hour,
                            'duration': duration,
                            'subject': subject,
                            'room': room,
                            'class': students_set,
                            'activity_id': act_id or None,
                            'source_file': os.path.basename(path),
                            'original_hour_field': hour_raw,
                            'original_day_field': day_raw,
                            'period': period  # إضافة معلومات الفترة
                        }

                        timetable_data.setdefault(teacher, []).append(activity)

                        # Also add to class timetable (فقط للأقسام الحقيقية)
                        if students_set:
                            main_class = extract_main_class(students_set)
                            if is_real_class(main_class):
                                class_activity = activity.copy()
                                class_activity['teacher'] = teacher
                                class_activity['original_class'] = students_set  # حفظ الاسم الأصلي
                                classes_timetable.setdefault(main_class, []).append(class_activity)
                                print(f"    ✅ إضافة نشاط للقسم: {main_class}")

                        total += 1
                        print(
                            f"    ✅ تم إضافة نشاط لـ {teacher}: {day_raw} الساعة {start_hour} - {subject} - {students_set}")

    except Exception as e:
        messagebox.showerror("Erreur import CSV", f"خطأ أثناء قراءة ملفات CSV:\n{e}")
        import traceback
        print(traceback.format_exc())
        return False

    # Finalize data structures
    for subject in materials_teachers:
        materials_teachers[subject] = sorted(list(materials_teachers[subject]))

    for teacher in teachers_subjects:
        teachers_subjects[teacher] = sorted(list(teachers_subjects[teacher]))

    for teacher in teachers_classes:
        teachers_classes[teacher] = sorted(list(teachers_classes[teacher]))

    for class_name in classes_teachers:
        classes_teachers[class_name] = sorted(list(classes_teachers[class_name]))

    # Assign colors
    mats = sorted(list(materials_teachers.keys()))
    for i, m in enumerate(mats):
        materials_colors[m] = DEFAULT_COLORS[i % len(DEFAULT_COLORS)]

    # Show import summary
    summary = f"✅ تم استيراد {total} نشاط من {len(paths)} ملف\n"
    summary += f"📚 المواد: {len(materials_teachers)} - 👨‍🏫 الأساتذة: {len(timetable_data)} - 🏫 الأقسام: {len(classes_timetable)}\n"

    if problematic_rows:
        summary += f"\n⚠️ تحذير: {len(problematic_rows)} صف به مشاكل في تحويل الأيام/الساعات\n"
        summary += "عينة من الصفوف المشكلة:\n"
        for prob in problematic_rows[:3]:
            summary += f"  - الصف {prob['row']}: اليوم '{prob['day_raw']}' -> {prob['day_converted']}, الساعة '{prob['hour_raw']}' -> {prob['hour_converted']}\n"

    messagebox.showinfo("استيراد ناجح", summary)

    # عرض عينة من البيانات المستوردة
    debug_timetable_data()
    return True


def debug_timetable_data():
    """عرض عينة من البيانات المستوردة للتdebug"""
    if not timetable_data:
        return

    debug_info = "📊 عينة من البيانات المستوردة:\n\n"

    # عرض عينة من الأساتذة
    debug_info += "👨‍🏫 عينة من جداول الأساتذة:\n"
    for teacher, activities in list(timetable_data.items())[:2]:
        debug_info += f"الأستاذ: {teacher}\n"
        for act in activities[:3]:
            day_map = {0: "الاثنين", 1: "الثلاثاء", 2: "الأربعاء", 3: "الخميس", 4: "الجمعة", 5: "السبت", 6: "الأحد"}
            day_name = day_map.get(act.get('weekday', ''), f"غير محدد ({act.get('original_day_field', '')})")
            start_hour = act.get('start_hour', 'غير محدد')
            period = act.get('period', 'غير محدد')
            debug_info += f"  - {day_name} الساعة {start_hour} ({period}): {act.get('subject', '')} - {act.get('class', '')}\n"
        debug_info += "\n"

    # عرض عينة من الأقسام
    debug_info += "🏫 عينة من جداول الأقسام:\n"
    for class_name, activities in list(classes_timetable.items())[:2]:
        debug_info += f"القسم: {class_name}\n"
        for act in activities[:3]:
            day_map = {0: "الاثنين", 1: "الثلاثاء", 2: "الأربعاء", 3: "الخميس", 4: "الجمعة", 5: "السبت", 6: "الأحد"}
            day_name = day_map.get(act.get('weekday', ''), "غير محدد")
            start_hour = act.get('start_hour', 'غير محدد')
            period = act.get('period', 'غير محدد')
            debug_info += f"  - {day_name} الساعة {start_hour} ({period}): {act.get('subject', '')} - {act.get('teacher', '')}\n"
        debug_info += "\n"

    print(debug_info)


# ---------------- دوال جديدة لعرض جداول الأقسام ----------------
def build_class_timetable(class_name):
    """بناء جدول القسم مع الساعات الصباحية والمسائية"""
    morning_hours = [8, 9, 10, 11]
    afternoon_hours = [14, 15, 16, 17]  # إصلاح: إضافة الساعة 17
    hours = [f"{h:02d}:00 - {h + 1:02d}:00" for h in morning_hours] + ["---"] + [f"{h:02d}:00 - {h + 1:02d}:00" for h in
                                                                                 afternoon_hours]

    days_order = [6, 0, 1, 2, 3]
    days_ar_map = {6: "الأحد", 0: "الاثنين", 1: "الثلاثاء", 2: "الأربعاء", 3: "الخميس"}

    header = ["الساعة / اليوم"] + [days_ar_map[d] for d in days_order]
    data = [header]

    for slot in hours:
        if slot == "---":
            data.append(["—"] + ["" for _ in days_order])
            continue
        row = [slot]
        slot_hour = int(slot.split(":")[0])
        for d in days_order:
            cell = ""
            for act in classes_timetable.get(class_name, []):
                if act.get('weekday') is None or act.get('start_hour') is None:
                    continue
                if act['weekday'] == d and act['start_hour'] == slot_hour:
                    subj = act.get('subject') or ""
                    teacher = act.get('teacher') or ""
                    room = act.get('room') or ""
                    if subj and teacher:
                        cell = f"{subj}\n{teacher}\n({room})" if room else f"{subj}\n{teacher}"
                    elif subj:
                        cell = subj
                    break
            row.append(cell)
        data.append(row)
    return data


def show_class_timetable_window(class_name):
    """عرض جدول القسم"""
    if class_name not in classes_timetable:
        messagebox.showinfo("معلومة", f"لا يوجد جدول مستورد لهذا القسم: {class_name}")
        return

    win = tk.Toplevel(root)
    win.title(f"جدول القسم - {class_name}")
    win.geometry("980x640")
    win.configure(bg=BG)

    tk.Label(win, text=f"🗓️ جدول القسم - {class_name}", font=("Arial", 14, "bold"), bg=BG).pack(pady=6)

    data = build_class_timetable(class_name)
    rows, cols = len(data), len(data[0])
    grid_frame = tk.Frame(win, bg=BG)
    grid_frame.pack(fill='both', expand=True, padx=8, pady=8)

    for i in range(rows):
        for j in range(cols):
            val = data[i][j]
            bg = "#ffffff"
            fg = "black"

            # header
            if i == 0:
                bg = "#263238"
                fg = "white"
            elif j == 0:
                # time column style
                if any(s in val for s in ("08:", "09:", "10:", "11:")):
                    bg = "#E8F5FF"
                elif any(s in val for s in ("14:", "15:", "16:", "17:", "18:")):
                    bg = "#FFF7E6"
                else:
                    bg = "#ECEFF1"
            if val.strip() == "—":
                bg = "#90A4AE"
                fg = "white"
            if i != 0 and j != 0 and val and val.strip() != "—":
                subj = val.split("\n")[0]
                bg = color_for_subject(subj)

            lbl = tk.Label(grid_frame, text=val, bg=bg, fg=fg, borderwidth=1, relief="solid",
                           width=18, height=3, justify="center", font=("Arial", 10), wraplength=140)
            # show RTL: place column reversed
            lbl.grid(row=i, column=(cols - j - 1), sticky="nsew", padx=1, pady=1)

    for c in range(cols):
        grid_frame.grid_columnconfigure(c, weight=1)
    for r in range(rows):
        grid_frame.grid_rowconfigure(r, weight=1)

    btn_frame = tk.Frame(win, bg=BG)
    btn_frame.pack(pady=6)

    ttk.Button(btn_frame, text="🖨️ طباعة / حفظ كـ PDF",
               command=lambda: export_class_pdf(class_name)).pack(side='left', padx=6)
    ttk.Button(btn_frame, text="👨‍🏫 الأساتذة المسندين",
               command=lambda: show_class_teachers(class_name)).pack(side='left', padx=6)


def export_class_pdf(class_name, pdf_path=None):
    """تصدير جدول القسم كـ PDF"""
    if not REPORTLAB_AVAILABLE:
        messagebox.showwarning("مكتبة مفقودة", "لتصدير PDF ثبّت reportlab (pip install reportlab)")
        return

    data = build_class_timetable(class_name)
    if pdf_path is None:
        fname = filedialog.asksaveasfilename(defaultextension=".pdf",
                                             filetypes=[("PDF files", "*.pdf")],
                                             title=f"حفظ جدول القسم {class_name}")
        if not fname:
            return
        pdf_path = fname

    try:
        page_w, page_h = landscape(A4)
        c = canvas.Canvas(pdf_path, pagesize=(page_w, page_h))
        c.setFont("Helvetica-Bold", 14)
        c.drawCentredString(page_w / 2, page_h - 30, f"جدول القسم - {class_name}")

        left = 40
        top = page_h - 70
        cols = len(data[0])
        rows = len(data)
        col_w = (page_w - 2 * left) / cols
        row_h = 24

        for r in range(rows):
            y = top - (r + 1) * row_h
            for ci in range(cols):
                x = left + ci * col_w
                c.rect(x, y, col_w, row_h, stroke=1, fill=0)
                text = data[r][ci]
                if r == 0:
                    c.setFont("Helvetica-Bold", 9)
                    c.drawCentredString(x + col_w / 2, y + 6, text)
                else:
                    c.setFont("Helvetica", 8)
                    if text:
                        lines = text.split("\n")
                        for li, line in enumerate(lines[:3]):  # عرض حتى 3 أسطر
                            c.drawString(x + 3, y + row_h - 8 - li * 8, line)

        c.save()
        messagebox.showinfo("تم", f"تم حفظ PDF في: {pdf_path}")

        try:
            if FITZ_AVAILABLE:
                os.startfile(pdf_path)
        except:
            pass

    except Exception as e:
        messagebox.showerror("خطأ PDF", f"خطأ أثناء توليد PDF:\n{e}")


def show_class_teachers(class_name):
    """عرض الأساتذة المسندين للقسم"""
    teachers = sorted(list(classes_teachers.get(class_name, [])))

    win = tk.Toplevel(root)
    win.title(f"أساتذة القسم - {class_name}")
    win.geometry("400x300")
    win.configure(bg=BG)

    tk.Label(win, text=f"👨‍🏫 أساتذة القسم {class_name}", font=("Arial", 14, "bold"), bg=BG).pack(pady=8)

    if not teachers:
        tk.Label(win, text="لا يوجد أساتذة مسندين لهذا القسم", bg=BG).pack(pady=8)
        return

    tree = ttk.Treeview(win, columns=("الأستاذ", "المواد"), show='headings', height=12)
    tree.heading("الأستاذ", text="الأستاذ")
    tree.heading("المواد", text="المواد")
    tree.column("الأستاذ", width=150)
    tree.column("المواد", width=200)
    tree.pack(fill='both', expand=True, padx=8, pady=8)

    for teacher in teachers:
        subjects = teachers_subjects.get(teacher, [])
        subjects_str = "، ".join(subjects[:3])  # عرض أول 3 مواد فقط
        if len(subjects) > 3:
            subjects_str += " ..."
        tree.insert('', 'end', values=(teacher, subjects_str))


# ---------------- دوال الواجهة المحسنة ----------------
def open_classes_window():
    """فتح نافذة عرض جميع الأقسام"""
    if not classes_timetable:
        messagebox.showinfo("معلومة", "لم يتم استيراد أي بيانات عن الأقسام بعد")
        return

    top = tk.Toplevel(root)
    top.title("جميع الأقسام")
    top.geometry("500x600")
    top.configure(bg=BG)

    tk.Label(top, text="🏫 جميع الأقسام", font=("Arial", 16, "bold"), bg=BG).pack(pady=12)

    # إنشاء إطار للقائمة وأزرار التحكم
    main_frame = tk.Frame(top, bg=BG)
    main_frame.pack(fill='both', expand=True, padx=10, pady=10)

    # شريط البحث
    search_frame = tk.Frame(main_frame, bg=BG)
    search_frame.pack(fill='x', pady=5)
    tk.Label(search_frame, text="بحث:", bg=BG).pack(side='left')
    search_var = tk.StringVar()
    search_entry = tk.Entry(search_frame, textvariable=search_var, width=30)
    search_entry.pack(side='left', padx=5)

    # قائمة الأقسام
    list_frame = tk.Frame(main_frame, bg=BG)
    list_frame.pack(fill='both', expand=True, pady=10)

    lb = tk.Listbox(list_frame, width=50, height=20, font=("Arial", 11))
    scrollbar = tk.Scrollbar(list_frame)
    lb.pack(side='left', fill='both', expand=True)
    scrollbar.pack(side='right', fill='y')
    lb.config(yscrollcommand=scrollbar.set)
    scrollbar.config(command=lb.yview)

    # تعبئة القائمة
    class_list = sorted(classes_timetable.keys())
    for class_name in class_list:
        teachers_count = len(classes_teachers.get(class_name, []))
        activities_count = len(classes_timetable.get(class_name, []))
        lb.insert('end', f"{class_name} ({teachers_count} أستاذ - {activities_count} حصة)")

    # أزرار التحكم
    btn_frame = tk.Frame(main_frame, bg=BG)
    btn_frame.pack(fill='x', pady=10)

    ttk.Button(btn_frame, text="📅 عرض جدول القسم",
               command=lambda: on_select_class(lb), width=20).pack(side='left', padx=5)
    ttk.Button(btn_frame, text="👨‍🏫 عرض أساتذة القسم",
               command=lambda: on_select_class_teachers(lb), width=20).pack(side='left', padx=5)

    def on_select_class(listbox):
        selection = listbox.curselection()
        if not selection:
            messagebox.showwarning("تحذير", "الرجاء اختيار قسم من القائمة")
            return
        class_full_text = listbox.get(selection[0])
        class_name = class_full_text.split(' (')[0]  # استخراج اسم القسم فقط
        show_class_timetable_window(class_name)

    def on_select_class_teachers(listbox):
        selection = listbox.curselection()
        if not selection:
            messagebox.showwarning("تحذير", "الرجاء اختيار قسم من القائمة")
            return
        class_full_text = listbox.get(selection[0])
        class_name = class_full_text.split(' (')[0]  # استخراج اسم القسم فقط
        show_class_teachers(class_name)

    # وظيفة البحث
    def update_list(event=None):
        search_text = search_var.get().lower()
        lb.delete(0, tk.END)
        for class_name in class_list:
            if search_text in class_name.lower():
                teachers_count = len(classes_teachers.get(class_name, []))
                activities_count = len(classes_timetable.get(class_name, []))
                lb.insert('end', f"{class_name} ({teachers_count} أستاذ - {activities_count} حصة)")

    search_var.trace('w', lambda name, index, mode: update_list())
    search_entry.bind('<Return>', update_list)


# ---------------- دوال التسجيل والمتابعة ----------------
def append_row_to_excel(date_str, prof, type_str, matiere, hour_str, note=""):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([date_str, prof, type_str, matiere, hour_str, note])
        wb.save(EXCEL_FILE)
    except Exception as e:
        messagebox.showerror("Erreur Excel", f"خطأ أثناء الكتابة في {EXCEL_FILE}:\n{e}")


def sessions_for_prof_on_date(prof, date_obj):
    if prof not in timetable_data:
        return None
    wd = date_obj.weekday()  # 0=Mon .. 6=Sun
    sessions = []
    for s in timetable_data.get(prof, []):
        s_wd = s.get('weekday')
        s_start = s.get('start_hour')
        s_dur = s.get('duration', 1)
        if s_start is None:
            continue
        if s_wd is not None and s_wd != wd:
            continue
        for h in range(s_start, s_start + max(1, s_dur)):
            if 8 <= h <= 20:
                sessions.append({'start_hour': h, 'subject': s.get('subject', ''), 'room': s.get('room', ''),
                                 'class': s.get('class', '')})
    uniq = {}
    for it in sessions:
        uniq[it['start_hour']] = it
    return [uniq[h] for h in sorted(uniq.keys())]


def open_daily_hours_window(prof, matiere):
    win = tk.Toplevel(root)
    win.title(f"تسجيل غيابات/تأخرات - {prof} - {matiere}")
    win.geometry("820x620");
    win.configure(bg=BG)
    tk.Label(win, text=f"تسجيل غيابات/تأخرات لـ {prof} ({matiere})", font=("Arial", 14, "bold"), bg=BG).pack(pady=8)
    topf = tk.Frame(win, bg=BG);
    topf.pack(pady=6)
    tk.Label(topf, text="📅 اختر التاريخ:", bg=BG).pack(side="left", padx=6)
    cal = DateEntry(topf, width=14, background='darkblue', foreground='white', borderwidth=2,
                    date_pattern='yyyy-mm-dd');
    cal.pack(side="left", padx=6)
    tk.Label(win, text="انقر على خانة الساعة لتغيير الحالة: 🟢 حاضر → 🔴 غياب → 🟠 تأخر → 🟢", bg=BG, fg="#333").pack(
        pady=6)
    all_hours = [f"{h:02d}:00 - {h + 1:02d}:00" for h in range(8, 17)]
    grid_frame = tk.Frame(win, bg=BG);
    grid_frame.pack(pady=8)
    btn_states = {};
    buttons = []

    def build_hour_grid_for_date(selected_date):
        for w in grid_frame.winfo_children(): w.destroy()
        sessions = sessions_for_prof_on_date(prof, selected_date)
        available = set()
        if sessions is not None:
            for s in sessions: available.add(s['start_hour'])
        cols = 3;
        r = 0;
        c = 0
        btn_states.clear();
        buttons.clear()
        for idx, label in enumerate(all_hours):
            hour_num = 8 + idx
            if sessions is not None and hour_num not in available:
                b = tk.Button(grid_frame, text=label + "\n(غير مقرر)", bg="#E5E8E8", width=24, height=3, wraplength=160,
                              state='disabled', relief="sunken")
            else:
                b = tk.Button(grid_frame, text=label, bg=COLOR_PRESENT, width=24, height=3, wraplength=160,
                              relief="raised", bd=2, font=("Arial", 10, "bold"))
                b.bind("<Button-1>", lambda e, idx=idx, btn=b: toggle_state(idx, btn))
                btn_states[idx] = 0
                buttons.append(b)
            b.grid(row=r, column=c, padx=10, pady=8)
            c += 1
            if c >= cols: c = 0; r += 1

    def toggle_state(idx, btn):
        st = btn_states.get(idx, 0);
        st = (st + 1) % 3;
        btn_states[idx] = st
        if st == 0:
            btn.config(bg=COLOR_PRESENT, text=all_hours[idx])
        elif st == 1:
            btn.config(bg=COLOR_ABSENT, text=all_hours[idx] + "\n(غياب)")
        else:
            btn.config(bg=COLOR_LATE, text=all_hours[idx] + "\n(تأخر)")

    build_hour_grid_for_date(datetime.date.today())

    def on_date_change(event=None):
        sel = cal.get_date();
        build_hour_grid_for_date(sel)

    cal.bind("<<DateEntrySelected>>", on_date_change)
    tk.Label(win, text="ملاحظة عامة (اختياري):", bg=BG).pack(pady=6)
    note_text = tk.Text(win, height=4, width=90);
    note_text.pack(pady=6)

    def save_today():
        date_str = cal.get_date().strftime('%Y-%m-%d');
        general_note = note_text.get('1.0', tk.END).strip();
        any_saved = False
        for idx, state in btn_states.items():
            if state == 0: continue
            hour_label = all_hours[idx];
            type_str = "غياب" if state == 1 else "تأخر"
            append_row_to_excel(date_str, prof, type_str, matiere, hour_label, general_note);
            any_saved = True
        if any_saved:
            messagebox.showinfo("تم الحفظ", f"تم حفظ الغيابات/التأخرات لـ {prof} بتاريخ {date_str}");
            win.destroy()
        else:
            messagebox.showinfo("لا شيء للتسجيل", "لم يتم اختيار أي ساعة (الكل حاضر).")

    ttk.Button(win, text="💾 حفظ اليوم", command=save_today).pack(pady=10)


# ---------------- دوال عرض جداول الأساتذة ----------------
def build_teacher_timetable(teacher):
    """بناء جدول الأستاذ مع الساعات الصباحية والمسائية"""
    morning_hours = [8, 9, 10, 11]
    afternoon_hours = [14, 15, 16, 17]  # إصلاح: إضافة الساعة 17
    hours = [f"{h:02d}:00 - {h + 1:02d}:00" for h in morning_hours] + ["---"] + [f"{h:02d}:00 - {h + 1:02d}:00" for h in
                                                                                 afternoon_hours]

    days_order = [6, 0, 1, 2, 3]
    days_ar_map = {6: "الأحد", 0: "الاثنين", 1: "الثلاثاء", 2: "الأربعاء", 3: "الخميس"}

    header = ["الساعة / اليوم"] + [days_ar_map[d] for d in days_order]
    data = [header]

    for slot in hours:
        if slot == "---":
            data.append(["—"] + ["" for _ in days_order])
            continue
        row = [slot]
        slot_hour = int(slot.split(":")[0])
        for d in days_order:
            cell = ""
            for act in timetable_data.get(teacher, []):
                if act.get('weekday') is None or act.get('start_hour') is None:
                    continue
                if act['weekday'] == d and act['start_hour'] == slot_hour:
                    subj = act.get('subject') or ""
                    clas = act.get('class') or ""
                    room = act.get('room') or ""
                    if subj:
                        cell = f"{subj}\n[{clas}] ({room})" if (clas or room) else subj
                    else:
                        cell = f"[{clas}] ({room})" if (clas or room) else ""
                    break
            row.append(cell)
        data.append(row)
    return data


def show_teacher_timetable_window(teacher):
    if teacher not in timetable_data:
        messagebox.showinfo("معلومة", f"لا يوجد emploi du temps مستورد لهذا الأستاذ: {teacher}")
        return
    win = tk.Toplevel(root);
    win.title(f"استعمال زمن - {teacher}");
    win.geometry("980x640");
    win.configure(bg=BG)
    tk.Label(win, text=f"🗓️ استخدام الزمن - {teacher}", font=("Arial", 14, "bold"), bg=BG).pack(pady=6)

    data = build_teacher_timetable(teacher)
    rows, cols = len(data), len(data[0])
    grid_frame = tk.Frame(win, bg=BG);
    grid_frame.pack(fill='both', expand=True, padx=8, pady=8)

    for i in range(rows):
        for j in range(cols):
            val = data[i][j]
            bg = "#ffffff";
            fg = "black"
            # header
            if i == 0:
                bg = "#263238";
                fg = "white"
            elif j == 0:
                # time column style
                if any(s in val for s in ("08:", "09:", "10:", "11:")):
                    bg = "#E8F5FF"
                elif any(s in val for s in ("14:", "15:", "16:", "17:", "18:")):
                    bg = "#FFF7E6"
                else:
                    bg = "#ECEFF1"
            if val.strip() == "—":
                bg = "#90A4AE";
                fg = "white"
            if i != 0 and j != 0 and val and val.strip() != "—":
                subj = val.split("\n")[0]
                bg = color_for_subject(subj)
            lbl = tk.Label(grid_frame, text=val, bg=bg, fg=fg, borderwidth=1, relief="solid",
                           width=18, height=3, justify="center", font=("Arial", 10), wraplength=140)
            # show RTL: place column reversed
            lbl.grid(row=i, column=(cols - j - 1), sticky="nsew", padx=1, pady=1)

    for c in range(cols):
        grid_frame.grid_columnconfigure(c, weight=1)
    for r in range(rows):
        grid_frame.grid_rowconfigure(r, weight=1)

    btn_frame = tk.Frame(win, bg=BG);
    btn_frame.pack(pady=6)
    ttk.Button(btn_frame, text="🖨️ طباعة / حفظ كـ PDF", command=lambda: export_teacher_pdf(teacher)).pack(side='left',
                                                                                                          padx=6)
    ttk.Button(btn_frame, text="🏫 الأقسام المسندة", command=lambda: show_assigned_classes(teacher)).pack(side='left',
                                                                                                         padx=6)


def export_teacher_pdf(teacher, pdf_path=None):
    if not REPORTLAB_AVAILABLE:
        messagebox.showwarning("مكتبة مفقودة", "لتصدير PDF ثبّت reportlab (pip install reportlab)")
        return
    data = build_teacher_timetable(teacher)
    if pdf_path is None:
        fname = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")],
                                             title="Enregistrer en PDF")
        if not fname:
            return
        pdf_path = fname
    try:
        page_w, page_h = landscape(A4)
        c = canvas.Canvas(pdf_path, pagesize=(page_w, page_h))
        c.setFont("Helvetica-Bold", 14)
        c.drawCentredString(page_w / 2, page_h - 30, f"جدول أسبوعي - {teacher}")
        left = 40;
        top = page_h - 70
        cols = len(data[0]);
        rows = len(data)
        col_w = (page_w - 2 * left) / cols;
        row_h = 24
        for r in range(rows):
            y = top - (r + 1) * row_h
            for ci in range(cols):
                x = left + ci * col_w
                c.rect(x, y, col_w, row_h, stroke=1, fill=0)
                text = data[r][ci]
                if r == 0:
                    c.setFont("Helvetica-Bold", 9);
                    c.drawCentredString(x + col_w / 2, y + 6, text)
                else:
                    c.setFont("Helvetica", 8)
                    if text:
                        lines = text.split("\n")
                        for li, line in enumerate(lines[:2]):
                            c.drawString(x + 3, y + row_h - 10 - li * 10, line)
        c.save()
        messagebox.showinfo("تم", f"تم حفظ PDF في: {pdf_path}")
        try:
            if FITZ_AVAILABLE:
                os.startfile(pdf_path)
        except:
            pass
    except Exception as e:
        messagebox.showerror("Erreur PDF", f"خطأ أثناء توليد PDF:\n{e}")


def show_assigned_classes(prof):
    classes = sorted(list(teachers_classes.get(prof, [])))
    subs = sorted(list(teachers_subjects.get(prof, [])))
    win = tk.Toplevel(root);
    win.title(f"الأقسام المسندة - {prof}");
    win.geometry("420x320");
    win.configure(bg=BG)
    tk.Label(win, text=f"الأقسام المسندة لـ {prof}", font=("Arial", 14, "bold"), bg=BG).pack(pady=8)
    tree = ttk.Treeview(win, columns=("القسم", "المادة"), show='headings', height=10)
    tree.heading("القسم", text="القسم");
    tree.heading("المادة", text="المادة")
    tree.column("القسم", width=160);
    tree.column("المادة", width=200)
    tree.pack(fill='both', expand=True, padx=8, pady=8)
    max_len = max(len(classes), len(subs))
    for i in range(max_len):
        cl = classes[i] if i < len(classes) else ""
        mat = subs[i] if i < len(subs) else ""
        tree.insert('', 'end', values=(cl, mat))
    if not classes:
        tk.Label(win, text="لا توجد أقسام مسندة مسجلة لهذا الأستاذ.", bg=BG).pack(pady=8)


# ---------------- دوال التقارير ----------------
def generate_pdf_for_prof(prof, periode, matiere=None, date_filter=None):
    if not REPORTLAB_AVAILABLE:
        messagebox.showwarning("تنبيه", "لتصدير PDF ثبّت reportlab (pip install reportlab)");
        return None
    wb = openpyxl.load_workbook(EXCEL_FILE);
    ws = wb.active;
    today = datetime.date.today()
    filename = os.path.join(REPORTS_DIR, f"{prof}_{periode}.pdf");
    c = canvas.Canvas(filename, pagesize=A4)
    c.setFont("Helvetica-Bold", 16);
    c.drawCentredString(300, 810, f"تقرير {periode} - {prof}")
    c.setFont("Helvetica", 10);
    c.drawString(50, 790, f"المادة: {matiere if matiere else 'جميع المواد'}");
    c.drawString(50, 775, f"تاريخ الطباعة: {today.strftime('%Y-%m-%d')}")
    y = 750
    for row in ws.iter_rows(min_row=2, values_only=True):
        date, row_prof, ttype, row_matiere, hour, note = row
        if row_prof != prof: continue
        if date_filter and date != date_filter: continue
        if periode == 'يومي' and date != today.strftime('%Y-%m-%d'): continue
        if periode == 'شهري' and not date.startswith(today.strftime('%Y-%m')): continue
        if periode == 'سنوي' and not date.startswith(str(today.year)): continue
        if matiere and row_matiere != matiere: continue
        text = f"{date} | {ttype} | {row_matiere} | {hour} | {note or ''}"
        max_len = 95
        while text:
            c.drawString(50, y, text[:max_len]);
            text = text[max_len:];
            y -= 12
            if y < 60:
                c.showPage();
                y = 800;
                c.setFont("Helvetica", 10)
    c.save();
    return filename


def preview_pdf_internal(pdf_file):
    if not FITZ_AVAILABLE:
        try:
            os.startfile(pdf_file)
        except Exception:
            messagebox.showinfo("ملف PDF", f"تم حفظ التقرير هنا: {pdf_file}")
        return
    preview = tk.Toplevel(root);
    preview.title("📄 معاينة التقرير");
    preview.geometry("900x700")
    canvas_frame = tk.Canvas(preview, bg="#f0f0f0");
    scroll_y = tk.Scrollbar(preview, orient="vertical", command=canvas_frame.yview)
    scrollable = tk.Frame(canvas_frame);
    scrollable.bind("<Configure>", lambda e: canvas_frame.configure(scrollregion=canvas_frame.bbox("all")))
    canvas_frame.create_window((0, 0), window=scrollable, anchor="nw");
    canvas_frame.configure(yscrollcommand=scroll_y.set)
    canvas_frame.pack(side="left", fill="both", expand=True);
    scroll_y.pack(side="right", fill="y")
    try:
        doc = fitz.open(pdf_file)
        for p in range(len(doc)):
            page = doc.load_page(p);
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
            img = Image.open(io.BytesIO(pix.tobytes("ppm")));
            photo = ImageTk.PhotoImage(img)
            lbl = tk.Label(scrollable, image=photo);
            lbl.image = photo;
            lbl.pack(pady=8)
    except Exception as e:
        tk.Label(scrollable, text=f"خطأ أثناء تحميل الملف: {e}", fg="red").pack(pady=10)


def show_report_table(prof, periode, matiere=None, date_filter=None):
    wb = openpyxl.load_workbook(EXCEL_FILE);
    ws = wb.active;
    today = datetime.date.today();
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        date, row_prof, ttype, row_matiere, hour, note = row
        if row_prof != prof: continue
        if date_filter and date != date_filter: continue
        if periode == 'يومي' and date != today.strftime('%Y-%m-%d'): continue
        if periode == 'شهري' and not date.startswith(today.strftime('%Y-%m')): continue
        if periode == 'سنوي' and not date.startswith(str(today.year)): continue
        if matiere and row_matiere != matiere: continue
        records.append((date, ttype, row_matiere, hour, note))
    win = tk.Toplevel(root);
    win.title(f"تقرير {periode} - {prof}");
    win.geometry("920x680")
    tk.Label(win, text=f"📋 تقرير {periode} - {prof}", font=("Arial", 16, "bold")).pack(pady=8)
    cols = ("Date", "Type", "Matière", "Heure", "Remarque");
    tree = ttk.Treeview(win, columns=cols, show='headings', height=18)
    for c in cols: tree.heading(c, text=c); tree.column(c, width=160 if c != 'Remarque' else 300)
    tree.pack(fill='both', expand=True, pady=6)
    for r in records: tree.insert('', 'end', values=r)

    def preview_and_save():
        pdf_file = generate_pdf_for_prof(prof, periode, matiere, date_filter)
        if pdf_file: preview_pdf_internal(pdf_file)

    ttk.Button(win, text='👁️ Afficher / Enregistrer en PDF', command=preview_and_save).pack(pady=8)


# ---------------- دوال الواجهات الفرعية ----------------
def open_material_window(matiere):
    top = tk.Toplevel(root);
    top.title(f"أساتذة {matiere}");
    top.geometry("380x480");
    top.configure(bg=BG)
    tk.Label(top, text=f"أساتذة {matiere}", font=("Arial", 14, "bold"), bg=BG).pack(pady=10)
    profs = materials_teachers.get(matiere, [])
    if not profs:
        tk.Label(top, text="لا يوجد أساتذة مسجلين لهذه المادة", bg=BG).pack(pady=8);
        return
    for p in profs:
        ttk.Button(top, text=p, width=34, command=lambda pr=p, m=matiere: open_prof_tracking_window(pr, m)).pack(pady=6)


def open_prof_tracking_window(prof, matiere):
    top = tk.Toplevel(root);
    top.title(f"ملف متابعة {prof}");
    top.geometry("560x520");
    top.configure(bg=BG)
    tk.Label(top, text=f"ملف متابعة {prof}", font=("Arial", 16, "bold"), bg=BG).pack(pady=12)
    ttk.Button(top, text="🕒 تسجيل غيابات/تأخر (جدول الساعات)", width=38,
               command=lambda: open_daily_hours_window(prof, matiere)).pack(pady=6)
    ttk.Button(top, text="📅 جدول الأستاذ", width=38, command=lambda: show_teacher_timetable_window(prof)).pack(pady=6)
    ttk.Button(top, text="🏫 الأقسام المسندة", width=38, command=lambda: show_assigned_classes(prof)).pack(pady=6)
    ttk.Button(top, text="📘 متابعة دفتر النصوص - تسجيل ملاحظة", width=38,
               command=lambda: open_text_note_window(prof, matiere, "دفتر النصوص")).pack(pady=6)
    ttk.Button(top, text="🚫 طرد التلاميذ - تسجيل", width=38,
               command=lambda: open_text_note_window(prof, matiere, "طرد")).pack(pady=6)
    ttk.Button(top, text="🧑‍🏫 مسؤول قسم/مادة - تسجيل", width=38,
               command=lambda: open_text_note_window(prof, matiere, "مسؤول")).pack(pady=6)
    tk.Label(top, text="التقارير:", bg=BG, font=("Arial", 12, "bold")).pack(pady=10)
    ttk.Button(top, text="🗓️ عرض تقرير ليوم محدد", width=34,
               command=lambda: show_report_for_date_window(prof, matiere)).pack(pady=4)
    ttk.Button(top, text="📅 تقرير شهري (هذا الشهر)", width=34,
               command=lambda: show_report_table(prof, 'شهري', matiere)).pack(pady=4)
    ttk.Button(top, text="📊 تقرير سنوي (هذا العام)", width=34,
               command=lambda: show_report_table(prof, 'سنوي', matiere)).pack(pady=4)


def open_text_note_window(prof, matiere, note_type):
    win = tk.Toplevel(root);
    win.title(f"{note_type} - {prof}");
    win.geometry("520x360");
    win.configure(bg=BG)
    tk.Label(win, text=f"{note_type} لـ {prof} ({matiere})", font=("Arial", 13, "bold"), bg=BG).pack(pady=8)
    tk.Label(win, text="📅 اختر التاريخ:", bg=BG).pack(pady=4)
    cal = DateEntry(win, width=14, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd');
    cal.pack(pady=4)
    tk.Label(win, text="الملاحظة:", bg=BG).pack(pady=6)
    text = tk.Text(win, height=10, width=70);
    text.pack(pady=6)

    def do_save():
        content = text.get('1.0', tk.END).strip()
        if not content:
            messagebox.showwarning("تنبيه", "اكتب الملاحظة قبل الحفظ");
            return
        date_str = cal.get_date().strftime('%Y-%m-%d')
        append_row_to_excel(date_str, prof, note_type, matiere, "", content)
        messagebox.showinfo("تم", f"تم حفظ {note_type} لـ {prof} بتاريخ {date_str}");
        win.destroy()

    ttk.Button(win, text="💾 حفظ", command=do_save).pack(pady=8)


def show_report_for_date_window(prof, matiere):
    win = tk.Toplevel(root);
    win.title("اختر التاريخ");
    win.geometry("320x180");
    win.configure(bg=BG)
    tk.Label(win, text=f"عرض تقرير ليوم محدد - {prof}", bg=BG).pack(pady=8)
    cal = DateEntry(win, width=14, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd');
    cal.pack(pady=8)

    def do_show():
        date_str = cal.get_date().strftime('%Y-%m-%d');
        show_report_table(prof, 'يومي', matiere, date_filter=date_str);
        win.destroy()

    ttk.Button(win, text="عرض", command=do_show).pack(pady=8)


def verify_timetable_match():
    """
    Verify that imported timetable matches the actual CSV structure
    """
    if not timetable_data:
        messagebox.showinfo("Info", "لم يتم استيراد أي بيانات بعد")
        return

    # Show summary of imported data
    summary = f"البيانات المستوردة:\n"
    summary += f"عدد الأساتذة: {len(timetable_data)}\n"
    summary += f"عدد المواد: {len(materials_teachers)}\n"
    summary += f"عدد الأقسام: {len(classes_timetable)}\n"
    summary += f"إجمالي الأنشطة: {sum(len(acts) for acts in timetable_data.values())}\n\n"

    # Show sample of first teacher's schedule
    if timetable_data:
        first_teacher = list(timetable_data.keys())[0]
        summary += f"عينة من جدول الأستاذ {first_teacher}:\n"
        for act in timetable_data[first_teacher][:3]:
            day_map = {0: "الاثنين", 1: "الثلاثاء", 2: "الأربعاء", 3: "الخميس", 4: "الجمعة", 5: "السبت", 6: "الأحد"}
            day_name = day_map.get(act.get('weekday', ''), f"غير محدد ({act.get('original_day_field', '')})")
            original_hour = act.get('original_hour_field', '')
            summary += f"- {day_name} الساعة {act.get('start_hour', '')} (الأصلي: {original_hour}): {act.get('subject', '')} - {act.get('class', '')}\n"

    # Show sample of classes
    if classes_timetable:
        first_class = list(classes_timetable.keys())[0]
        summary += f"\nعينة من جدول القسم {first_class}:\n"
        for act in classes_timetable[first_class][:3]:
            day_map = {0: "الاثنين", 1: "الثلاثاء", 2: "الأربعاء", 3: "الخميس", 4: "الجمعة", 5: "السبت", 6: "الأحد"}
            day_name = day_map.get(act.get('weekday', ''), "غير محدد")
            summary += f"- {day_name} الساعة {act.get('start_hour', '')}: {act.get('subject', '')} - {act.get('teacher', '')}\n"

    messagebox.showinfo("تحقق من البيانات", summary)


# ---------------- دوال الاستيراد الأساسية ----------------
def import_fet_activities_csv(path):
    """Wrapper to import a single path"""
    return import_fet_activities_csv_files([path] if path else [])


def try_auto_import_sample():
    """Try to auto-import known candidate CSVs, else scan all CSVs for FET-like headers"""
    existing = [p for p in CSV_CANDIDATES if os.path.exists(p)]
    if existing:
        return import_fet_activities_csv_files(existing)
    candidates = []
    for fn in os.listdir('.'):
        if not fn.lower().endswith('.csv'):
            continue
        try:
            with open(fn, encoding='utf-8-sig') as f:
                head = f.readline()
                if 'Activity' in head or 'Teachers' in head or 'Subject' in head:
                    candidates.append(fn)
        except:
            continue
    if candidates:
        return import_fet_activities_csv_files(sorted(candidates))
    return False


def import_csv_and_refresh():
    """Open file dialog to select multiple CSVs and import, then refresh UI."""
    file_paths = filedialog.askopenfilenames(title="استيراد جدول CSV من FET",
                                             filetypes=[("CSV files", "*.csv"), ("All files", "*.*")])
    if not file_paths:
        return
    ok = import_fet_activities_csv_files(list(file_paths))
    if ok:
        build_main_ui()


# ---------------- الواجهة الرئيسية ----------------
def build_main_ui():
    for w in root.winfo_children():
        w.destroy()

    tk.Label(root, text='📚 برنامج منصوري لمتابعة عمل الأساتذة', font=("Arial", 22, "bold"), bg=BG, fg='#004d80').pack(
        pady=8)

    top_controls = tk.Frame(root, bg=BG)
    top_controls.pack(pady=6)

    ttk.Button(top_controls, text="📅 استيراد جدول CSV", command=import_csv_and_refresh).pack(side='left', padx=8)
    ttk.Button(top_controls, text="🔍 التحقق من البيانات", command=verify_timetable_match).pack(side='left', padx=8)
    ttk.Button(top_controls, text="🏫 عرض جميع الأقسام", command=open_classes_window).pack(side='left', padx=8)

    tk.Label(top_controls, text="(استيراد CSV من FET)", bg=BG).pack(side='left', padx=8)

    stats_frame = tk.Frame(root, bg=BG)
    stats_frame.pack(pady=6)

    tk.Label(stats_frame, text=f"📚 المواد: {len(materials_teachers)}", bg=BG, font=("Arial", 11)).pack(side='left',
                                                                                                       padx=12)
    tk.Label(stats_frame, text=f"👨‍🏫 الأساتذة: {len(timetable_data)}", bg=BG, font=("Arial", 11)).pack(side='left',
                                                                                                       padx=12)
    tk.Label(stats_frame, text=f"🏫 الأقسام: {len(classes_timetable)}", bg=BG, font=("Arial", 11)).pack(side='left',
                                                                                                       padx=12)

    main_frame = tk.Frame(root, bg=BG)
    main_frame.pack(pady=10, fill='both', expand=True)

    # left: grid of subjects
    left = tk.Frame(main_frame, bg=BG)
    left.pack(side='left', fill='both', expand=True, padx=8, pady=8)

    if not materials_teachers:
        tk.Label(left, text="لم يتم استيراد الجدول بعد. اضغط 'استيراد جدول CSV' أو ضع ملف CSV في المجلد.", bg=BG).pack(
            pady=12)
    else:
        mats = list(materials_teachers.keys())
        cols = 3
        r = 0
        c = 0
        for i, mat in enumerate(mats):
            color = materials_colors.get(mat, DEFAULT_COLORS[i % len(DEFAULT_COLORS)])
            btn = tk.Button(left, text=mat, bg=color, font=("Arial", 12, "bold"), fg="black", width=26, height=2,
                            relief="raised", bd=2, command=lambda m=mat: open_material_window(m))
            btn.grid(row=r, column=c, padx=10, pady=10)
            c += 1
            if c >= cols:
                c = 0
                r += 1

    # right: teachers list and actions
    right = tk.Frame(main_frame, bg=BG, width=360)
    right.pack(side='right', fill='y', padx=8, pady=8)

    tk.Label(right, text="👨‍🏫 قائمة الأساتذة:", font=("Arial", 12, "bold"), bg=BG).pack(pady=6)

    lb = tk.Listbox(right, width=36, height=20)
    lb.pack(pady=4)

    for t in sorted(timetable_data.keys()):
        lb.insert('end', t)

    def on_select_teacher(evt=None):
        sel = lb.curselection()
        if not sel:
            return
        prof = lb.get(sel[0])
        open_prof_tracking_window(prof, (teachers_subjects.get(prof, [""])[0] if teachers_subjects.get(prof) else ""))

    lb.bind("<Double-Button-1>", on_select_teacher)

    ttk.Button(right, text="فتح ملف الأستاذ", command=on_select_teacher).pack(pady=6)
    ttk.Button(right, text="استيراد CSV", command=import_csv_and_refresh).pack(pady=6)
    ttk.Button(right, text="عرض الأقسام", command=open_classes_window).pack(pady=6)

    if not FITZ_AVAILABLE:
        tk.Label(root, text='ملاحظة: لم يتم تثبيت PyMuPDF/Pillow للمعاينة الداخلية PDF. (pip install PyMuPDF pillow)',
                 fg='red', bg=BG).pack(side='bottom', pady=6)


# ---------------- تشغيل التطبيق ----------------
root = tk.Tk()
root.title('ناظر المدرسة - Suivi des enseignants')
root.geometry('1120x760')
root.configure(bg=BG)

# try auto import then build UI
if try_auto_import_sample():
    build_main_ui()
else:
    tk.Label(root, text='📚 برنامج الناظر لمتابعة عمل الأساتذة', font=("Arial", 22, "bold"), bg=BG, fg='#004d80').pack(
        pady=8)
    ttk.Button(root, text="📅 استيراد جدول CSV", command=import_csv_and_refresh).pack(pady=12)
    tk.Label(root, text="ضع ملفات FET CSV في نفس المجلد أو اضغط 'استيراد جدول CSV' لاختيارها.", bg=BG).pack(pady=6)

root.mainloop()
